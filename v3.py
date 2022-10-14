# %%
import imp
import openpyxl
from openpyxl.utils.cell import get_column_letter, column_index_from_string
from collections import defaultdict
from copy import copy
from openpyxl.styles import PatternFill
import os

%%
all_path = [r"a"]
start = 'shb.xlsx'
PATH_TEMPLATE = '2.xlsx'
end = 'fn.xlsx'
for path in all_path:
    start_file = os.listdir(path)[0]
    start = path + '\\' + start_file
    end = path +'\\' + PREFIX_END + start_file



# %%
    start = 'shb.xlsx'
    PATH_TEMPLATE = '2.xlsx'
    end = 'fn.xlsx'

    import re

    def color_all(sheat, sorce_color_column): 
        for x in range (1,sheat.max_row+1):
            color_row(sheat,x, sorce_color_column)

        return sheat
    def color_row(sheat, index, sorce_color_column):
        for x in range (1,sheat.max_column+1):
            sheat.cell(row=index, column =x).fill= copy(sheat.cell(row=index, column =sorce_color_column).fill)
        return sheat
    def copy_row(sheat, index ):
        sheat.insert_rows(index)
        for x in range (1,sheat.max_column+1):
            sheat.cell(row=index, column =x).value = sheat.cell(row=index+1, column =x).value
            sheat.cell(row=index, column =x).fill= copy(sheat.cell(row=index+1, column =x).fill)

        return sheat

    def table_to_str(a):
        text = ""
        for x in a:
            text += x + ", " 
        return text[:-2]

    def def_val(value : str):

        value = value.replace('application-default', '')
        def to_dict(m):
            dd =defaultdict(list)
        
        m = re.findall(r'[0-9]+-[a-z]{3}-[a-z|0-9]+', value)
        if m :
            return ('OTHER', m[0].replace('OTHER:', '').replace('other:', '') )

        elif re.findall(r'(OTHER|other):.[A-z]+',value):
            a = re.findall(r'[A-z]+', value)[1].replace('application', '')
            return ('OTHER', a)

        elif re.findall(r'[A-z]+.[0-9]+-[0-9]+',value):
            nr = re.findall(r'[A-z]+', value)[0]
            po = re.findall(r'[0-9]+-[0-9]+', value)[0]
            return (nr,po)
            
        elif re.findall(r'[0-9]+-[0-9]+.[A-z]+',value):
            nr = re.findall(r'[A-z]+', value)[0]
            po = re.findall(r'[0-9]+-[0-9]+', value)[0]
            return (nr,po)

        elif re.findall(r'([0-9]+-[0-9]+)', value) :
            return ('LAST', re.findall(r'[0-9]+-[0-9]+', value)[0])

        elif re.findall(r'([0-9]+.[A-z]+)', value) :
            nr = re.findall(r'[A-z]+', value)[0]
            po = re.findall(r'[0-9]+', value)[0]
            return (nr,po)

        elif re.findall(r'([A-z]+.[0-9]+)', value) :
            nr = re.findall(r'[A-z]+', value)[0]
            po = re.findall(r'[0-9]+', value)[0]
            return (nr,po)
        
        elif re.findall(r'[A-z|-]{2,}', value) :
            return ('OTHER', re.findall(r'[A-z|-]{2,}', value)[0])

        elif re.findall(r'[A-z]+', value) :
            return ('OTHER', re.findall(r'[A-z]+', value)[0])

        elif re.findall(r'[0-9]+', value) :
            return ('LAST', re.findall(r'[0-9]+', value)[0])

    # %%
    wb = openpyxl.load_workbook(start)
    sh = wb.get_sheet_by_name(wb.get_sheet_names()[0])

    # %%
    wb2 = openpyxl.load_workbook(PATH_TEMPLATE)
    sh2 = wb2.get_sheet_by_name(wb2.get_sheet_names()[0])


    # %%
    SORCE_START_ROW = 6
    OUTPUT_START_ROW = 6

    COLUMN_DICT = {'C':'C',
                    'E':'E',
                    'J':'J',
                    'L':'L',
                    'M' : 'M'}
                    
    JOIN_CCOLUMNS = (('M','N'),  'Q')
    PROTOCOLS = ('M', ('M','N'))

    # %%
    # copy 


                    
    for r_iny, r_out in COLUMN_DICT.items():
        iny = column_index_from_string(r_iny)
        out = column_index_from_string(r_out)
        temp = []
        for x in range(SORCE_START_ROW +1, sh.max_row +1):
        
            temp.append((sh.cell(column=iny, row=x).value, copy(sh.cell(column=iny, row=x).fill) ))



        for x in range(OUTPUT_START_ROW +1 , OUTPUT_START_ROW +1 + len(temp)):
            sh2.cell(column=out, row=x).value = temp[0][0]
            sh2.cell(column=out, row=x).fill = temp[0][1]

            del temp[0]


    temp = []
    prot = column_index_from_string(JOIN_CCOLUMNS[0][0])
    service = column_index_from_string(JOIN_CCOLUMNS[0][1])
    coment = column_index_from_string(JOIN_CCOLUMNS[1] )

    for x in range(SORCE_START_ROW +1, sh.max_row +1):
        def if_none(v):
            if v == None:
                return ""
            else :
                return v
        temp.append(f" {if_none(sh.cell(column=prot, row=x).value)} {if_none(sh.cell(column=service, row=x).value)}")

    for x in range(OUTPUT_START_ROW +1 , OUTPUT_START_ROW +1 + len(temp)):
        sh2.cell(column=coment, row=x).value = temp[0]
        del temp[0]







    # %%
    def spliter(row):
            V = row.replace(" ","").split(',')
            dd =defaultdict(list)
            final = []
            for i in V :
                if def_val(i)[0] == "LAST":
                    dd[ret[0].lower()].append(def_val(i)[1])
                else :
                    ret = def_val(i)
                    dd[ret[0].lower()].append(ret[1])
            
            
            all_values =list(dd.values())
            keysy = list(dd.keys())
        
            for keys, values in zip(keysy, all_values):
                
                if (keys == 'other') and (len(values) > 1)  :
                    for valu in values:
                        protocol = keys
                        port =  valu
                        final.append((protocol, port))

                    
                # elif len(keys) > 1 :
                #     for keu in dd.keys():
                #         protocol = keu
                #         port =  table_to_str(dd[keu])
                #         final.append((protocol, port))

                else :
                    protocol = keys
                    port = table_to_str(values)
                    final.append((protocol, port))
                
            return final

    def spliter_pipe(sheat, col):

        for i in range (OUTPUT_START_ROW,sheat.max_row+1):
            if sheat.cell(row=i, column =col).value :
                piped = sheat.cell(row=i, column =col).value.replace(" ","").split('|')
                
                print('aa', piped)
                if len(piped) > 1:
                    # copy cell ilosc splita -1 
                    for _ in range(len(piped) -1):
                        copy_row(sheat, i)

                    for ep, ele in enumerate(piped):
                        sheat.cell(row=i+ep, column =col).value = ele

        return sheat
        

    # %%

    sorce = column_index_from_string(PROTOCOLS[0])
    protocols_out = column_index_from_string(PROTOCOLS[1][0])
    port_out = column_index_from_string(PROTOCOLS[1][1] )

    sh2 = spliter_pipe(sh2, sorce)

    x = SORCE_START_ROW +1

    while sh2.cell(column=3, row=x).value:
        
        v = sh2.cell(column=sorce, row=x).value
        
        if v :


            splited = spliter(v)
            #print(splited)

            if len(splited) ==1:
                sh2.cell(column=protocols_out, row=x).value = splited[0][0]
                sh2.cell(column=port_out, row=x).value =splited[0][1]
                x +=1
            else :

                
                for i , (in_pr, in_por) in enumerate(splited):
                    if in_pr == 'other' and len(in_por) > 1:
                        pass
                    sh2.cell(column=protocols_out, row=x).value = in_pr
                    sh2.cell(column=port_out, row=x).value = in_por
                    if i+1 < len(splited):
                        sh2 =copy_row(sh2 , x)

                    x +=1

        else :
            sh2.cell(column=protocols_out, row=x).value = ""
            sh2.cell(column=port_out, row=x).value = ""
            x +=1

    # %%
    sh2 = color_all(sh2,3)
    wb2.save(end)

    print(f'excel {end} was generatet')



